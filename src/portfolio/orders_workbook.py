"""Build RotationOrders.xlsx — the manual-months order sheet.

One command regenerates the workbook from the latest ledger row:

    python -m src.portfolio.orders_workbook

The user then opens it, picks the risk tier, types the account value and
current holdings, and reads off dollar amounts to trade (IBKR order
entry accepts dollar amounts directly with fractional shares enabled).
The workbook embeds the CURRENT month's resolved weights — regenerate it
each month after the ledger row lands; the screener's output IS the
Data sheet. Interim tool while the automated executor is staged
(AUTOMATION.md); the executor replaces it.
"""

from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from src.portfolio.matrix import MATRIX_VERSION, TIERS
from src.portfolio.paper_logger import LEDGER_PATH, load_ledger

OUT_PATH = Path(__file__).resolve().parents[2] / "RotationOrders.xlsx"

ARIAL = "Arial"
YELLOW = PatternFill("solid", fgColor="FFFF00")
GREY = PatternFill("solid", fgColor="F2F2F2")
NAVY = PatternFill("solid", fgColor="1F3864")
BLUE = Font(name=ARIAL, size=10, color="0000FF")
GREEN = Font(name=ARIAL, size=10, color="008000")
BOX = Border(*[Side(style="thin", color="BFBFBF")] * 4)
CUR = "$#,##0"
CUR_SIGNED = "$#,##0;($#,##0);-"

BOOK_ROWS = 8      # output rows 12..19
HOLD_ROWS = 10     # holdings rows 12..21
BLOCK = 10         # Data-sheet rows per tier block


def _font(size=10, bold=False, color="000000"):
    return Font(name=ARIAL, size=size, bold=bold, color=color)


def build(ledger_path: Path = LEDGER_PATH, out_path: Path = OUT_PATH) -> Path:
    ledger = load_ledger(ledger_path)
    if ledger.empty:
        raise SystemExit("ledger is empty — run the paper logger first")
    row = ledger.iloc[-1]
    month, quadrant = str(row["month"]), str(row["quadrant"])
    allocs: dict[str, dict[str, float]] = json.loads(row["allocations"])

    wb = Workbook()

    # ---------------- Data sheet: this month's resolved weights ----------
    data = wb.create_sheet("Data")
    data["H1"], data["H2"], data["H3"] = month, quadrant, str(row["matrix_version"])
    data["G1"], data["G2"], data["G3"] = "Month", "Quadrant", "Matrix"
    for cells in (("G1", "G2", "G3"),):
        for c in cells:
            data[c].font = _font(9, bold=True)
    for i, tier in enumerate(TIERS):
        base = i * BLOCK + 1
        data.cell(row=base, column=1, value=tier).font = _font(10, bold=True)
        for j, (ticker, weight) in enumerate(allocs[tier].items(), start=1):
            data.cell(row=base + j, column=1, value=ticker).font = _font(10)
            c = data.cell(row=base + j, column=2, value=round(weight, 4))
            c.number_format = "0.00%"
    data.sheet_properties.tabColor = "808080"

    # ---------------- Orders sheet ---------------------------------------
    ws = wb.active
    ws.title = "Orders"
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:I1")
    ws["A1"] = "MONTHLY ORDER SHEET — QUADRANT ROTATION"
    ws["A1"].font = _font(13, bold=True, color="FFFFFF")
    ws["A1"].fill = NAVY
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A2:I2")
    ws["A2"] = ('="Ledger row: "&Data!$H$1&"  ·  "&Data!$H$2&"  ·  matrix "&Data!$H$3&'
                '"  —  regenerate monthly: python -m src.portfolio.orders_workbook"')
    ws["A2"].font = GREEN
    ws.merge_cells("A3:I3")
    ws["A3"] = ("YELLOW cells are yours: pick the tier, enter total account value, list "
                "current holdings (example rows below show the format — overwrite the 0s). "
                "Everything else computes. Trade the SELL ALL rows first, then the book.")
    ws["A3"].font = _font(9, color="595959")

    ws["A5"] = "Risk tier"
    ws["A5"].font = _font(10, bold=True)
    ws["C5"] = "VAGG"
    dv = DataValidation(type="list", formula1='"CONS,MOD,AGG,VAGG"', allow_blank=False)
    ws.add_data_validation(dv)
    dv.add(ws["C5"])
    ws["A6"] = "Total account value ($)"
    ws["A6"].font = _font(10, bold=True)
    ws["C6"] = 11000  # user input: cash + market value of rotation positions
    ws["C6"].number_format = CUR
    ws["A7"] = "= settled cash + market value of current rotation positions"
    ws["A7"].font = _font(8, color="595959")
    for c in ("C5", "C6"):
        ws[c].fill = YELLOW
        ws[c].font = BLUE
        ws[c].border = BOX
        ws[c].alignment = Alignment(horizontal="center")
    ws["K5"] = "tier idx"
    ws["K5"].font = _font(8, color="BFBFBF")
    ws["L5"] = '=MATCH($C$5,{"CONS";"MOD";"AGG";"VAGG"},0)'
    ws["L5"].font = _font(8, color="BFBFBF")

    # holdings input table (left)
    ws["A10"] = "Your current holdings"
    ws["A10"].font = _font(10, bold=True)
    for col, header in (("A", "Ticker"), ("B", "Value ($)"), ("C", "Status")):
        c = ws[f"{col}11"]
        c.value = header
        c.font = _font(9, bold=True)
        c.fill = GREY
        c.border = BOX
    prefill = list(allocs["VAGG"])  # example rows: this month's book at $0
    for k in range(HOLD_ROWS):
        r = 12 + k
        t = ws.cell(row=r, column=1, value=prefill[k] if k < len(prefill) else None)
        v = ws.cell(row=r, column=2, value=0 if k < len(prefill) else None)
        for c in (t, v):
            c.fill = YELLOW
            c.font = BLUE
            c.border = BOX
        v.number_format = CUR
        s = ws.cell(row=r, column=3)
        s.value = (f'=IF($A{r}="","",IF(COUNTIF($D$12:$D$19,$A{r})>0,'
                   f'"in book","SELL ALL"))')
        s.font = _font(9)
        s.border = BOX

    # target book table (right)
    ws["D10"] = "This month's target book"
    ws["D10"].font = _font(10, bold=True)
    headers = ["Ticker", "Weight", "Target $", "Held $", "Action", "Trade $ (+buy / −sell)"]
    for j, header in enumerate(headers):
        c = ws.cell(row=11, column=4 + j, value=header)
        c.font = _font(9, bold=True)
        c.fill = GREY
        c.border = BOX
    for k in range(BOOK_ROWS):
        r = 12 + k
        idx = f"{k + 1}+($L$5-1)*{BLOCK}+1"
        ws.cell(row=r, column=4).value = (
            f'=IF(INDEX(Data!$A:$A,{idx})=0,"",INDEX(Data!$A:$A,{idx}))')
        ws.cell(row=r, column=5).value = f'=IF($D{r}="","",INDEX(Data!$B:$B,{idx}))'
        ws.cell(row=r, column=5).number_format = "0.00%"
        ws.cell(row=r, column=6).value = f'=IF($D{r}="","",ROUND(E{r}*$C$6,0))'
        ws.cell(row=r, column=6).number_format = CUR
        ws.cell(row=r, column=7).value = (
            f'=IF($D{r}="","",SUMIF($A$12:$A$21,$D{r},$B$12:$B$21))')
        ws.cell(row=r, column=7).number_format = CUR
        ws.cell(row=r, column=8).value = (
            f'=IF($D{r}="","",IF(F{r}-G{r}>0,"BUY",'
            f'IF((G{r}-F{r})/$C$6>0.05,"SELL","HOLD")))')
        ws.cell(row=r, column=9).value = (
            f'=IF($D{r}="","",IF(H{r}="BUY",F{r}-G{r},'
            f'IF(H{r}="SELL",-(G{r}-F{r}),0)))')
        ws.cell(row=r, column=9).number_format = CUR_SIGNED
        for col in range(4, 10):
            ws.cell(row=r, column=col).border = BOX
            ws.cell(row=r, column=col).font = _font(10)

    ws["D20"] = "Totals"
    ws["D20"].font = _font(10, bold=True)
    ws["F20"] = "=SUM(F12:F19)"
    ws["F20"].number_format = CUR
    ws["I20"] = "=SUM(I12:I19)"
    ws["I20"].number_format = CUR_SIGNED
    ws["A23"] = "Off-book positions to SELL entirely (left table, SELL ALL rows)"
    ws["A23"].font = _font(10)
    ws["C23"] = '=SUMPRODUCT(($C$12:$C$21="SELL ALL")*$B$12:$B$21)'
    ws["C23"].number_format = CUR
    ws["A24"] = "Cash remainder after all trades"
    ws["A24"].font = _font(10)
    ws["C24"] = "=$C$6-SUM(F12:F19)"
    ws["C24"].number_format = CUR
    ws["A26"] = ("Execution: sells first (SELL ALL rows, then book SELLs), then buys. Limit "
                 "orders near the midpoint; IBKR order entry accepts these dollar amounts "
                 "directly ('Amount in USD') with fractional shares on. Equity ETFs near the "
                 "close; never miss a rotation waiting for the right hour (DEPLOYMENT.md).")
    ws["A26"].font = _font(8, color="595959")
    ws["A27"] = ("The 5pp drift band is built in: overweights inside the band read HOLD — "
                 "do not sell them. HOLD/BUY logic matches src/portfolio/order_calc.py, "
                 "which the automated executor also uses.")
    ws["A27"].font = _font(8, color="595959")

    widths = {"A": 34, "B": 12, "C": 12, "D": 10, "E": 9, "F": 11, "G": 11, "H": 9, "I": 20,
              "K": 8, "L": 6}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for i in range(1, 10):
        data.column_dimensions[get_column_letter(i)].width = 12

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


if __name__ == "__main__":
    path = build()
    print(f"wrote {path}")
